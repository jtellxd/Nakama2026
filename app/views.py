from django.shortcuts import render, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_http_methods
from .models import Empleado, TipoAsistencia, RegistroAsistencia, DispositivoEmpleado
from .services import AsistenciaService, ReporteService
from .qr_service import QRService
from .utils import obtener_fecha_hora_actual
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import timedelta
from django.utils import timezone
import json


def es_staff(user):
    """
    Verifica si el usuario es staff (administrador).
    
    Args:
        user: Usuario a verificar
        
    Returns:
        bool: True si es staff y está autenticado
    """
    return user.is_authenticated and user.is_staff

@user_passes_test(es_staff)
def pagina_descarga_excel(request):
    """
    Página para descargar reportes de Excel.
    Solo accesible para usuarios staff.
    """
    return render(request, 'pagina_descarga_excel.html')


@user_passes_test(es_staff)
def exportar_resumen_excel(request):
    """
    Exporta un resumen de asistencia en formato Excel.
    Una hoja por empleado con su información, mostrando todos los días del rango
    (incluso los días sin registros, con la fecha siempre presente).
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Eliminar hoja vacía por defecto

    # Estilos
    titulo_font = Font(bold=True, color="FFFFFF", size=12)
    titulo_fill = PatternFill("solid", fgColor="1F3864")
    info_label_font = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    sin_registro_fill = PatternFill("solid", fgColor="F2F2F2")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Rango de fechas: desde el primer registro hasta hoy
    primera_fecha = RegistroAsistencia.objects.order_by('fecha_registro') \
        .values_list('fecha_registro', flat=True).first()
    hoy = timezone.localtime().date()

    if primera_fecha:
        dias_total = (hoy - primera_fecha).days + 1
        todas_las_fechas = [primera_fecha + timedelta(days=i) for i in range(dias_total)]
    else:
        todas_las_fechas = [hoy]

    # Obtener todos los registros agrupados
    datos_diarios = ReporteService.obtener_datos_resumen()

    # Obtener todos los empleados
    empleados = Empleado.objects.order_by('apellidos', 'nombres')

    if not empleados.exists():
        ws = wb.create_sheet("Sin Datos")
        ws.append(["No hay empleados registrados."])
    else:
        for empleado in empleados:
            # Nombre de la hoja (máx 31 chars, sin caracteres inválidos para Excel)
            chars_invalidos = ['/', '\\', '?', '*', '[', ']', ':']
            nombre_hoja = f"{empleado.apellidos}, {empleado.nombres}"
            for c in chars_invalidos:
                nombre_hoja = nombre_hoja.replace(c, '-')
            nombre_hoja = nombre_hoja[:31]

            ws = wb.create_sheet(title=nombre_hoja)

            # --- Fila 1: Título ---
            ws.append(["REPORTE DE ASISTENCIA"])
            ws.merge_cells('A1:E1')
            ws['A1'].font = titulo_font
            ws['A1'].fill = titulo_fill
            ws['A1'].alignment = Alignment(horizontal='center')

            # --- Filas 2-5: Información del empleado ---
            ws.append(["Nombre completo:", empleado.nombre_completo])
            ws['A2'].font = info_label_font
            ws.append(["Apellidos:", empleado.apellidos])
            ws['A3'].font = info_label_font
            ws.append(["Nombres:", empleado.nombres])
            ws['A4'].font = info_label_font
            ws.append(["Código QR:", empleado.codigo_qr or "No asignado"])
            ws['A5'].font = info_label_font

            # --- Fila 6: Espacio en blanco ---
            ws.append([])

            # --- Fila 7: Encabezados de columnas ---
            encabezados = [
                "Fecha", "Tiempo de Almuerzo",
                "Horas por Comisión", "Horas por Permiso (Otros)",
                "Horas Trabajadas Totales"
            ]
            ws.append(encabezados)
            fila_encabezado = ws.max_row
            for cell in ws[fila_encabezado]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            # --- Filas de datos: un día por fila ---
            for fecha in todas_las_fechas:
                key = (empleado.id_empleado, fecha)
                if key in datos_diarios:
                    data = datos_diarios[key]
                    horas = ReporteService.calcular_horas_empleado(data)
                    ws.append([
                        fecha.strftime("%Y-%m-%d"),
                        horas['almuerzo'],
                        horas['comision'],
                        horas['permiso'],
                        horas['trabajadas']
                    ])
                else:
                    # Día sin registro: fecha presente, resto vacío
                    ws.append([
                        fecha.strftime("%Y-%m-%d"),
                        "", "", "", ""
                    ])
                    for cell in ws[ws.max_row]:
                        cell.fill = sin_registro_fill

            # Ajustar ancho de columnas
            for i, col in enumerate(ws.columns, 1):
                col_letter = get_column_letter(i)
                max_length = max(
                    (len(str(cell.value)) for cell in col
                     if cell.value and cell.data_type != 'n' or (cell.value and str(cell.value))),
                    default=10
                )
                ws.column_dimensions[col_letter].width = max(max_length + 2, 20)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=resumen_asistencia.xlsx'
    wb.save(response)
    return response

@user_passes_test(es_staff)
def exportar_asistencia_excel(request):
    """
    Exporta registros de asistencia en formato Excel.
    Una hoja por empleado con su información, mostrando todos los días del rango.
    Columnas: Fecha + una columna por cada tipo de asistencia (hora registrada).
    Días sin registro: fecha presente, resto vacío con fondo gris.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Estilos
    titulo_font  = Font(bold=True, color="FFFFFF", size=12)
    titulo_fill  = PatternFill("solid", fgColor="1F3864")
    info_font    = Font(bold=True)
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="4F81BD")
    sin_reg_fill = PatternFill("solid", fgColor="F2F2F2")
    thin_border  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Orden lógico de tipos de asistencia
    orden_tipos = [
        'Entrada', 'Inicio Almuerzo', 'Fin Almuerzo', 'Salida',
        'Entrada por comisión', 'Salida por comisión',
        'Entrada por otros', 'Salida por otros',
    ]
    tipos_db = {t.nombre_asistencia for t in TipoAsistencia.objects.all()}
    tipos_cols = [n for n in orden_tipos if n in tipos_db]
    for nombre in tipos_db:
        if nombre not in tipos_cols:
            tipos_cols.append(nombre)

    num_cols = 1 + len(tipos_cols)  # Fecha + tipos
    col_fin = get_column_letter(num_cols)

    # Rango de fechas: primer registro → hoy
    primera_fecha = RegistroAsistencia.objects.order_by('fecha_registro') \
        .values_list('fecha_registro', flat=True).first()
    hoy = timezone.localtime().date()
    if primera_fecha:
        dias_total = (hoy - primera_fecha).days + 1
        todas_las_fechas = [primera_fecha + timedelta(days=i) for i in range(dias_total)]
    else:
        todas_las_fechas = [hoy]

    # Lookup rápido: {(empleado_id, fecha): {tipo_nombre: "HH:MM"}}
    lookup = {}
    for reg in RegistroAsistencia.objects.select_related('empleado', 'tipo') \
            .order_by('fecha_registro', 'hora_registro'):
        key = (reg.empleado.id_empleado, reg.fecha_registro)
        if key not in lookup:
            lookup[key] = {}
        lookup[key][reg.tipo.nombre_asistencia] = reg.hora_registro.strftime('%H:%M')

    empleados = Empleado.objects.order_by('apellidos', 'nombres')

    if not empleados.exists():
        ws = wb.create_sheet("Sin Datos")
        ws.append(["No hay empleados registrados."])
    else:
        for empleado in empleados:
            chars_invalidos = ['/', '\\', '?', '*', '[', ']', ':']
            nombre_hoja = f"{empleado.apellidos}, {empleado.nombres}"
            for c in chars_invalidos:
                nombre_hoja = nombre_hoja.replace(c, '-')
            nombre_hoja = nombre_hoja[:31]

            ws = wb.create_sheet(title=nombre_hoja)

            # Fila 1: Título
            ws.append(["REGISTRO DE ASISTENCIA"])
            ws.merge_cells(f'A1:{col_fin}1')
            ws['A1'].font = titulo_font
            ws['A1'].fill = titulo_fill
            ws['A1'].alignment = Alignment(horizontal='center')

            # Filas 2-5: Info del empleado
            ws.append(["Nombre completo:", empleado.nombre_completo])
            ws['A2'].font = info_font
            ws.append(["Apellidos:", empleado.apellidos])
            ws['A3'].font = info_font
            ws.append(["Nombres:", empleado.nombres])
            ws['A4'].font = info_font
            ws.append(["Código QR:", empleado.codigo_qr or "No asignado"])
            ws['A5'].font = info_font

            # Fila 6: Espacio en blanco
            ws.append([])

            # Fila 7: Encabezados de columnas
            encabezados = ["Fecha"] + tipos_cols
            ws.append(encabezados)
            fila_enc = ws.max_row
            for cell in ws[fila_enc]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            # Filas de datos: un día por fila
            for fecha in todas_las_fechas:
                key = (empleado.id_empleado, fecha)
                if key in lookup:
                    data = lookup[key]
                    fila = [fecha.strftime("%d/%m/%Y")]
                    for tipo in tipos_cols:
                        fila.append(data.get(tipo, ""))
                    ws.append(fila)
                else:
                    # Día sin registro: fecha + celdas vacías en gris
                    ws.append([fecha.strftime("%d/%m/%Y")] + [""] * len(tipos_cols))
                    for cell in ws[ws.max_row]:
                        cell.fill = sin_reg_fill

            # Ajustar ancho de columnas
            for i, col in enumerate(ws.columns, 1):
                col_letter = get_column_letter(i)
                max_length = max(
                    (len(str(cell.value)) for cell in col if cell.value),
                    default=10
                )
                ws.column_dimensions[col_letter].width = max(max_length + 2, 18)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=registro_asistencia.xlsx'
    wb.save(response)
    return response



def pagina_principal(request):
    """
    Página principal con opciones de acceso.
    """
    return render(request, 'pagina_principal.html')


def escanear_qr(request):
    """
    Página para escanear código QR.
    """
    return render(request, 'escanear_qr.html')


@ensure_csrf_cookie
def identificar_dispositivo(request):
    """
    Página de QR general: identifica por fingerprint. Si ya está vinculado, redirige directo al formulario.
    Si no, muestra selector de empleado para vincular el dispositivo.
    """
    empleados = Empleado.objects.order_by('apellidos', 'nombres')
    return render(request, 'identificar.html', { 'empleados': empleados })


def registrar_asistencia_qr(request, codigo_qr):
    """
    Vista para registrar asistencia usando código QR.
    Detecta automáticamente al empleado.
    """
    empleado = Empleado.buscar_por_codigo_qr(codigo_qr)
    
    if not empleado:
        messages.error(request, 'Código QR no válido o empleado no encontrado.')
        return render(request, 'error_qr.html')
    
    tipos_evento = TipoAsistencia.objects.all()

    if request.method == 'POST':
        tipo_id = request.POST.get('tipo_evento')
        descripcion = request.POST.get('descripcion') or ''
        fingerprint = request.POST.get('fingerprint')

        # Validar datos requeridos
        if not tipo_id:
            messages.error(request, 'Debe seleccionar un tipo de asistencia.')
            return render(request, 'formulario_qr.html', {
                'empleado': empleado,
                'tipos_evento': tipos_evento
            })

        # Usar el servicio para crear el registro
        success, message, registro = AsistenciaService.crear_registro_asistencia(
            empleado.id_empleado, tipo_id, descripcion, fingerprint
        )

        if success:
            messages.success(request, message)
            fecha, hora = obtener_fecha_hora_actual()
            return render(request, 'asistencia_exitosa.html', {
                'fecha': fecha,
                'hora': hora,
                'empleado': registro.empleado
            })
        messages.error(request, message)

    return render(request, 'formulario_qr.html', {
        'empleado': empleado,
        'tipos_evento': tipos_evento
    })


def registrar_asistencia_auto(request, empleado_id):
    """
    Registro usando identificación automática por fingerprint (QR general).
    Primera vez: se vincula en identificar_dispositivo.
    """
    empleado = get_object_or_404(Empleado, id_empleado=empleado_id)
    tipos_evento = TipoAsistencia.objects.all()

    if request.method == 'POST':
        tipo_id = request.POST.get('tipo_evento')
        descripcion = request.POST.get('descripcion') or ''
        fingerprint = request.POST.get('fingerprint')

        if not tipo_id:
            messages.error(request, 'Debe seleccionar un tipo de asistencia.')
            return render(request, 'formulario_qr.html', {
                'empleado': empleado,
                'tipos_evento': tipos_evento
            })

        success, message, registro = AsistenciaService.crear_registro_asistencia(
            empleado.id_empleado, tipo_id, descripcion, fingerprint
        )

        if success:
            messages.success(request, message)
            fecha, hora = obtener_fecha_hora_actual()
            return render(request, 'asistencia_exitosa.html', {
                'fecha': fecha,
                'hora': hora,
                'empleado': registro.empleado
            })
        messages.error(request, message)

    return render(request, 'formulario_qr.html', {
        'empleado': empleado,
        'tipos_evento': tipos_evento
    })


@require_http_methods(["POST", "OPTIONS"])
def api_buscar_empleado_qr(request):
    """
    API para buscar empleado por código QR.
    """
    # Responder preflight/local OPTIONS
    if request.method == 'OPTIONS':
        return JsonResponse({'success': True})
    try:
        data = json.loads(request.body)
        codigo_qr = data.get('codigo_qr')
        
        if not codigo_qr:
            return JsonResponse({'success': False, 'error': 'Código QR requerido'}, status=400)
        
        resultado = QRService.buscar_empleado_por_qr(codigo_qr)
        status_code = 200 if resultado.get('success') else 404
        return JsonResponse(resultado, status=status_code)
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error del servidor: {str(e)}'}, status=500)


@require_http_methods(["POST", "OPTIONS"])
def api_identificar_por_fingerprint(request):
    """
    Identifica empleado por fingerprint del dispositivo.
    """
    if request.method == 'OPTIONS':
        return JsonResponse({'success': True})
    try:
        data = json.loads(request.body)
        fingerprint = data.get('fingerprint')
        if not fingerprint:
            return JsonResponse({'success': False, 'error': 'Fingerprint requerido'}, status=400)
        empleado = DispositivoEmpleado.obtener_empleado_por_fingerprint(fingerprint)
        if empleado:
            return JsonResponse({
                'success': True,
                'empleado': {
                    'id': empleado.id_empleado,
                    'nombres': empleado.nombres,
                    'apellidos': empleado.apellidos,
                    'nombre_completo': empleado.nombre_completo,
                }
            })
        else:
            return JsonResponse({'success': False, 'error': 'Dispositivo no vinculado a un empleado'}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error del servidor: {str(e)}'}, status=500)


@require_http_methods(["POST", "OPTIONS"])
def api_vincular_fingerprint(request):
    """
    Vincula el fingerprint al empleado seleccionado (primera vez).
    """
    if request.method == 'OPTIONS':
        return JsonResponse({'success': True})
    try:
        data = json.loads(request.body)
        empleado_id = data.get('empleado_id')
        fingerprint = data.get('fingerprint')
        if not empleado_id or not fingerprint:
            return JsonResponse({'success': False, 'error': 'Empleado y fingerprint requeridos'}, status=400)
        empleado = Empleado.objects.get(id_empleado=empleado_id)
        # Reasignación permitida: si el fingerprint existe con otro empleado, se actualiza al elegido
        DispositivoEmpleado.objects.update_or_create(
            fingerprint=fingerprint,
            defaults={'empleado': empleado}
        )
        return JsonResponse({'success': True, 'empleado_id': empleado.id_empleado}, status=201)
    except Empleado.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Empleado no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error del servidor: {str(e)}'}, status=500)


@require_http_methods(["POST", "OPTIONS"])
def api_desvincular_fingerprint(request):
    """
    Desvincula el fingerprint del dispositivo actual para permitir seleccionar de nuevo.
    """
    if request.method == 'OPTIONS':
        return JsonResponse({'success': True})
    try:
        data = json.loads(request.body)
        fingerprint = data.get('fingerprint')
        if not fingerprint:
            return JsonResponse({'success': False, 'error': 'Fingerprint requerido'}, status=400)
        borrados, detalle = DispositivoEmpleado.objects.filter(fingerprint=fingerprint).delete()
        return JsonResponse({'success': True, 'deleted': borrados})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error del servidor: {str(e)}'}, status=500)


def registrar_asistencia(request):
    """
    Vista tradicional para registrar la asistencia de un empleado.
    Mantenida para compatibilidad.
    """
    empleados = Empleado.objects.all()
    tipos_evento = TipoAsistencia.objects.all()

    if request.method == 'POST':
        empleado_id = request.POST.get('empleado')
        tipo_id = request.POST.get('tipo_evento')
        descripcion = request.POST.get('descripcion') or ''
        fingerprint = request.POST.get('fingerprint')

        # Validar datos requeridos
        if not empleado_id or not tipo_id:
            messages.error(request, 'Debe seleccionar un empleado y tipo de asistencia.')
            return render(request, 'formulario.html', {
                'empleados': empleados,
                'tipos_evento': tipos_evento
            })

        # Usar el servicio para crear el registro
        success, message, registro = AsistenciaService.crear_registro_asistencia(
            empleado_id, tipo_id, descripcion, fingerprint
        )

        if success:
            messages.success(request, message)
            fecha, hora = obtener_fecha_hora_actual()
            return render(request, 'asistencia_exitosa.html', {
                'fecha': fecha,
                'hora': hora,
                'empleado': registro.empleado
            })
        messages.error(request, message)

    return render(request, 'formulario.html', {
        'empleados': empleados,
        'tipos_evento': tipos_evento
    })
